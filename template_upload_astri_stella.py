import streamlit as st
import openpyxl
import gdown
import os
import io

# Judul Web
st.title("Aplikasi Pemetaan Data Excel")

# 1. Download TEMPLATE dari Google Drive
file_id = '1SV5YYaIRsubz-6nsin5F-6yWj64-Ljjn'
template_file = "TEMPLATE_UPLOAD.xlsx"

if not os.path.exists(template_file):
    gdown.download(f'https://drive.google.com/uc?id={file_id}', template_file, quiet=False)

# 2. Upload File di Streamlit
uploaded_file = st.file_uploader("Upload file Excel data awal Anda:", type=['xlsx'])

if uploaded_file:
    if st.button("Proses Data"):
        wb_source = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = wb_source.sheetnames
        
        for sheet_name in sheet_names:
            ws_source = wb_source[sheet_name]
            wb_temp = openpyxl.load_workbook(template_file)
            ws_temp = wb_temp["CLUSTER"]

            # Pemetaan Sel Tunggal
            ws_temp["A2"] = ws_source["Y10"].value
            ws_temp["B2"] = ws_source["AR10"].value
            ws_temp["C2"] = ws_source["Z10"].value
            ws_temp["F2"] = ws_source["N10"].value
            ws_temp["I2"] = ws_source["O10"].value
            ws_temp["L2"] = ws_source["Q10"].value
            ws_temp["M2"] = ws_source["P10"].value

            # Pemetaan Kolom Banyak Baris
            col_mapping = {'AD': 'G', 'AE': 'H', 'AF': 'I', 'AG': 'J', 'AH': 'K', 'AI': 'L', 'AJ': 'M', 'J': 'U', 'K': 'V', 'AS': 'W', 'AU': 'Y', 'AV': 'Z'}
            
            tgt_row = 6
            for r in range(10, ws_source.max_row + 1):
                # Cek apakah baris kosong
                if all(ws_source[f"{src_col}{r}"].value is None for src_col in col_mapping.keys()):
                    continue
                for src_col, tgt_col in col_mapping.items():
                    ws_temp[f"{tgt_col}{tgt_row}"].value = ws_source[f"{src_col}{r}"].value
                tgt_row += 1

            # Simpan ke memori untuk di-download
            output = io.BytesIO()
            wb_temp.save(output)
            
            # Tombol Download per sheet
            st.download_button(
                label=f"Download Hasil: {sheet_name}",
                data=output.getvalue(),
                file_name=f"UPDATED_{sheet_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.success("Proses Selesai!")
